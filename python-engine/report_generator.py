#!/usr/bin/env python3
"""
Report Generation Module for Fire Department Analytics
Generates PDF and Excel reports
"""
import json
from datetime import datetime
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.enums import TA_CENTER, TA_JUSTIFY
from reportlab.lib import colors
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from analytics import FireDeptAnalytics

class ReportGenerator:
    def __init__(self):
        self.analytics = FireDeptAnalytics()
        # Use relative path
        import os
        script_dir = os.path.dirname(os.path.abspath(__file__))
        self.reports_dir = os.path.join(script_dir, '..', 'reports')
    
    def generate_executive_report_pdf(self, filters=None):
        """Generate comprehensive executive PDF report"""
        filename = f"{self.reports_dir}/Executive_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
        
        doc = SimpleDocTemplate(filename, pagesize=letter,
                               rightMargin=72, leftMargin=72,
                               topMargin=72, bottomMargin=48)
        
        styles = getSampleStyleSheet()
        story = []
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=TA_CENTER,
            fontName='Helvetica-Bold'
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12,
            fontName='Helvetica-Bold'
        )
        
        # Title
        story.append(Paragraph("Fire Department Analytics", title_style))
        story.append(Paragraph("Executive Summary Report", styles['Heading2']))
        story.append(Spacer(1, 0.3*inch))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}", styles['Normal']))
        story.append(Spacer(1, 0.5*inch))
        
        # KPIs Section
        story.append(Paragraph("Key Performance Indicators", heading_style))
        kpis = self.analytics.get_dashboard_kpis()
        
        kpi_data = [
            ['Metric', 'Value'],
            ['Total Incidents', f"{kpis['total_incidents']:,}"],
            ['Recent Incidents (30 days)', f"{kpis['recent_incidents_30d']:,}"],
            ['Average Response Time', f"{kpis['avg_response_time']} minutes"],
            ['Total Casualties', f"{kpis['total_casualties']:,}"],
            ['Total Injuries', f"{kpis['total_injuries']:,}"],
            ['Critical Incidents', f"{kpis['critical_incidents']:,}"],
            ['Active Fire Stations', f"{kpis['active_stations']:,}"],
            ['Property Damage', f"₹{kpis['total_property_damage']:,}"]
        ]
        
        kpi_table = Table(kpi_data, colWidths=[3*inch, 2*inch])
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        story.append(kpi_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Incidents by Type
        story.append(Paragraph("Incident Distribution by Type", heading_style))
        by_type = self.analytics.get_incidents_by_type()
        type_data = [['Incident Type', 'Count', 'Percentage']]
        total = sum(item['count'] for item in by_type)
        for item in by_type[:8]:
            percentage = (item['count'] / total * 100)
            type_data.append([item['type'], str(item['count']), f"{percentage:.1f}%"])
        
        type_table = Table(type_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
        type_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        story.append(type_table)
        story.append(PageBreak())
        
        # State-wise Analysis
        story.append(Paragraph("State-wise Analysis", heading_style))
        by_state = self.analytics.get_incidents_by_state()
        state_data = [['State', 'Incidents', 'Casualties', 'Avg Response Time']]
        for item in by_state[:10]:
            state_data.append([
                item['state'], 
                str(item['incidents']), 
                str(item['casualties']),
                f"{item['avg_response_time']} min"
            ])
        
        state_table = Table(state_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1.5*inch])
        state_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        story.append(state_table)
        story.append(Spacer(1, 0.3*inch))
        
        # Response Time Analysis
        story.append(Paragraph("Response Time Analysis", heading_style))
        response_analysis = self.analytics.get_response_time_analysis()
        response_data = [['Severity', 'Avg Time', 'Min Time', 'Max Time', 'Count']]
        for item in response_analysis:
            response_data.append([
                item['severity'],
                f"{item['avg_time']} min",
                f"{item['min_time']} min",
                f"{item['max_time']} min",
                str(item['count'])
            ])
        
        response_table = Table(response_data, colWidths=[1.5*inch, 1.3*inch, 1.3*inch, 1.3*inch, 1.1*inch])
        response_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        story.append(response_table)
        story.append(PageBreak())
        
        # Vendor Performance
        story.append(Paragraph("Top Vendor Performance", heading_style))
        vendors = self.analytics.get_vendor_performance()
        vendor_data = [['Vendor Name', 'Deliveries', 'On-Time %', 'Quality', 'Defects']]
        for vendor in vendors[:10]:
            vendor_data.append([
                vendor['vendor_name'][:30],
                str(vendor['total_deliveries']),
                f"{vendor['on_time_percentage']:.1f}%",
                f"{vendor['avg_quality_score']:.2f}",
                str(vendor['total_defects'])
            ])
        
        vendor_table = Table(vendor_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch, 1*inch])
        vendor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2c3e50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        story.append(vendor_table)
        
        # Build PDF
        doc.build(story)
        return filename
    
    def generate_excel_report(self, filters=None):
        """Generate comprehensive Excel report with charts"""
        filename = f"{self.reports_dir}/Analytics_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # 1. Dashboard Sheet
        ws_dashboard = wb.create_sheet("Dashboard")
        self._create_dashboard_sheet(ws_dashboard)
        
        # 2. Incidents Sheet
        ws_incidents = wb.create_sheet("Incidents Analysis")
        self._create_incidents_sheet(ws_incidents)
        
        # 3. State Analysis Sheet
        ws_states = wb.create_sheet("State Analysis")
        self._create_state_sheet(ws_states)
        
        # 4. Vendor Performance Sheet
        ws_vendors = wb.create_sheet("Vendor Performance")
        self._create_vendor_sheet(ws_vendors)
        
        # 5. Monthly Trends Sheet
        ws_trends = wb.create_sheet("Monthly Trends")
        self._create_trends_sheet(ws_trends)
        
        # 6. Raw Data Sheet
        ws_raw = wb.create_sheet("Raw Data")
        self._create_raw_data_sheet(ws_raw)
        
        wb.save(filename)
        return filename
    
    def _create_dashboard_sheet(self, ws):
        """Create dashboard KPIs sheet"""
        kpis = self.analytics.get_dashboard_kpis()
        
        # Title
        ws['A1'] = "Fire Department Analytics Dashboard"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Date
        ws['A2'] = f"Generated: {datetime.now().strftime('%B %d, %Y %I:%M %p')}"
        ws.merge_cells('A2:D2')
        
        # Headers
        ws['A4'] = "Key Performance Indicator"
        ws['B4'] = "Value"
        ws['A4'].font = Font(bold=True, color="FFFFFF")
        ws['B4'].font = Font(bold=True, color="FFFFFF")
        ws['A4'].fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        ws['B4'].fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        
        # KPIs
        row = 5
        kpi_labels = {
            'total_incidents': 'Total Incidents',
            'recent_incidents_30d': 'Recent Incidents (30 days)',
            'avg_response_time': 'Average Response Time (min)',
            'total_casualties': 'Total Casualties',
            'total_injuries': 'Total Injuries',
            'critical_incidents': 'Critical Incidents',
            'active_stations': 'Active Fire Stations',
            'active_vendors': 'Active Vendors',
            'total_property_damage': 'Total Property Damage (₹)'
        }
        
        for key, label in kpi_labels.items():
            ws[f'A{row}'] = label
            ws[f'B{row}'] = kpis[key]
            ws[f'B{row}'].number_format = '#,##0' if 'damage' not in key else '₹#,##0'
            row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
    
    def _create_incidents_sheet(self, ws):
        """Create incidents analysis sheet"""
        # Headers
        ws['A1'] = "Incident Type Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:D1')
        
        ws['A3'] = "Incident Type"
        ws['B3'] = "Count"
        ws['C3'] = "Percentage"
        for cell in ['A3', 'B3', 'C3']:
            ws[cell].font = Font(bold=True, color="FFFFFF")
            ws[cell].fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        
        # Data
        by_type = self.analytics.get_incidents_by_type()
        total = sum(item['count'] for item in by_type)
        
        row = 4
        for item in by_type:
            ws[f'A{row}'] = item['type']
            ws[f'B{row}'] = item['count']
            ws[f'C{row}'] = f"{item['count']/total*100:.1f}%"
            row += 1
        
        # Add chart
        chart = PieChart()
        chart.title = "Incident Distribution by Type"
        labels = Reference(ws, min_col=1, min_row=4, max_row=row-1)
        data = Reference(ws, min_col=2, min_row=3, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        ws.add_chart(chart, "E3")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def _create_state_sheet(self, ws):
        """Create state analysis sheet"""
        ws['A1'] = "State-wise Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:E1')
        
        # Headers
        headers = ['State', 'Incidents', 'Casualties', 'Avg Response Time (min)']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        
        # Data
        by_state = self.analytics.get_incidents_by_state()
        row = 4
        for item in by_state:
            ws[f'A{row}'] = item['state']
            ws[f'B{row}'] = item['incidents']
            ws[f'C{row}'] = item['casualties']
            ws[f'D{row}'] = item['avg_response_time']
            row += 1
        
        # Add bar chart
        chart = BarChart()
        chart.title = "Incidents by State"
        chart.y_axis.title = "Number of Incidents"
        
        data = Reference(ws, min_col=2, min_row=3, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "F3")
        
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
    
    def _create_vendor_sheet(self, ws):
        """Create vendor performance sheet"""
        ws['A1'] = "Vendor Performance Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:F1')
        
        headers = ['Vendor Name', 'Type', 'Deliveries', 'On-Time %', 'Quality Score', 'Defects']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        
        vendors = self.analytics.get_vendor_performance()
        row = 4
        for vendor in vendors:
            ws[f'A{row}'] = vendor['vendor_name']
            ws[f'B{row}'] = vendor['vendor_type']
            ws[f'C{row}'] = vendor['total_deliveries']
            ws[f'D{row}'] = vendor['on_time_percentage']
            ws[f'E{row}'] = vendor['avg_quality_score']
            ws[f'F{row}'] = vendor['total_defects']
            row += 1
        
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25
        for col in ['C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 15
    
    def _create_trends_sheet(self, ws):
        """Create monthly trends sheet"""
        ws['A1'] = "Monthly Trends"
        ws['A1'].font = Font(size=14, bold=True)
        ws.merge_cells('A1:C1')
        
        headers = ['Month', 'Incidents', 'Casualties']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        
        trends = self.analytics.get_monthly_trends()
        row = 4
        for item in trends:
            ws[f'A{row}'] = item['month']
            ws[f'B{row}'] = item['incidents']
            ws[f'C{row}'] = item['casualties']
            row += 1
        
        # Add line chart
        chart = LineChart()
        chart.title = "Monthly Incident Trends"
        chart.y_axis.title = "Count"
        
        data = Reference(ws, min_col=2, min_row=3, max_col=3, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row-1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, "E3")
        
        for col in ['A', 'B', 'C']:
            ws.column_dimensions[col].width = 18
    
    def _create_raw_data_sheet(self, ws):
        """Create raw incident data sheet"""
        ws['A1'] = "Raw Incident Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        headers = ['Incident ID', 'Date', 'State', 'City', 'Type', 'Severity', 
                  'Response Time', 'Casualties', 'Injuries', 'Property Damage']
        
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
        
        row = 4
        for incident in self.analytics.incidents[:100]:  # Limit to 100 for demo
            ws[f'A{row}'] = incident['incident_id']
            ws[f'B{row}'] = incident['timestamp']
            ws[f'C{row}'] = incident['state']
            ws[f'D{row}'] = incident['city']
            ws[f'E{row}'] = incident['incident_type']
            ws[f'F{row}'] = incident['severity']
            ws[f'G{row}'] = incident['response_time_minutes']
            ws[f'H{row}'] = incident['casualties']
            ws[f'I{row}'] = incident['injuries']
            ws[f'J{row}'] = incident['property_damage_inr']
            row += 1
        
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 18

def main():
    """Test report generation"""
    print("Generating reports...")
    generator = ReportGenerator()
    
    # Generate PDF
    pdf_file = generator.generate_executive_report_pdf()
    print(f"✓ PDF Report: {pdf_file}")
    
    # Generate Excel
    excel_file = generator.generate_excel_report()
    print(f"✓ Excel Report: {excel_file}")
    
    print("\n✅ Reports generated successfully!")

if __name__ == "__main__":
    main()
